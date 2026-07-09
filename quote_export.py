"""
견적서 출력 (Track B) - 업체용/소비자용 2종, Excel + PDF

- 소비자용(대외용): 제안가 기준 (최종고객 제출용)
- 업체용(파트너용): 공급가 기준 (유통 파트너 공급가)
공급가/마진 노출 분리, VR 중복방지 문구 포함.
"""
import os
import datetime

COMPANY = {
    "name": "솔루텍㈜",
    "addr": "서울 금천구 가산디지털1로 128 STX-V타워",
    "tel": "02-2169-0000",
    "fax": "02-2169-0099",
    "home": "http://www.solu.co.kr",
}

# (시트명, 단가키, 금액키, 단가헤더)
VARIANTS = {
    "소비자용": ("제안단가", "제안금액", "단가(제안)"),
    "업체용": ("공급단가", "공급금액", "단가(공급)"),
}


def _meta_rows(customer, company, kind):
    today = datetime.date.today().strftime("%Y-%m-%d")
    return [
        ["수신", customer.get("고객사명", ""), "회사명", COMPANY["name"]],
        ["사업명", customer.get("건명", ""), "견적일자", today],
        ["담당자", customer.get("담당자명", ""), "유효기간", "견적일로부터 30일"],
        ["견적담당", company.get("견적담당자명", ""), "구분", f"{kind} 견적서"],
    ]


def export_excel(path, customer, company, lines, note=None):
    """소비자용/업체용 2개 시트를 한 xlsx로 저장. openpyxl 필요."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="FFD700")
    bold = Font(bold=True)

    for kind, (unit_key, amt_key, unit_hdr) in VARIANTS.items():
        ws = wb.create_sheet(kind)
        ws.append(["QUOTATION", "", "", ""])
        ws["A1"].font = Font(bold=True, size=14)
        for r in _meta_rows(customer, company, kind):
            ws.append(r)
        ws.append([])
        hdr = ["No", "품목명", "조달상태", "단위", "수량", unit_hdr, "금액"]
        ws.append(hdr)
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")
        total = 0
        for i, l in enumerate(lines, 1):
            ws.append([i, l["품목명"], l["조달상태"], l["단위"],
                       l["수량"], l[unit_key], l[amt_key]])
            total += l[amt_key]
        ws.append(["", "", "", "", "", "합계(VAT별도)", total])
        ws.cell(row=ws.max_row, column=6).font = bold
        ws.cell(row=ws.max_row, column=7).font = bold
        if note:
            ws.append([])
            ws.append(["특이사항", note])
        # 열 폭
        widths = [5, 38, 12, 8, 6, 14, 14]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = w

    wb.save(path)
    return path


def export_pdf(path, customer, company, lines, kind="소비자용", note=None):
    """단일 종류(기본 소비자용) PDF. fpdf + arialuni.TTF 필요(레포 동봉)."""
    from fpdf import FPDF
    unit_key, amt_key, unit_hdr = VARIANTS[kind]

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("ArialUnicode", "", "arialuni.ttf", uni=True)
    pdf.set_font("ArialUnicode", size=14)
    pdf.cell(0, 10, txt=f"견적서 ({kind})", ln=True, align="C")
    pdf.ln(4)
    pdf.set_font("ArialUnicode", size=9)
    pdf.cell(0, 7, txt=f"수신: {customer.get('고객사명','')}  |  사업명: {customer.get('건명','')}", ln=True)
    pdf.cell(0, 7, txt=f"견적일자: {datetime.date.today()}  |  {COMPANY['name']}", ln=True)
    pdf.ln(3)

    widths = [10, 70, 22, 14, 14, 30, 30]
    headers = ["No", "품목명", "조달상태", "단위", "수량", unit_hdr, "금액"]
    for w, h in zip(widths, headers):
        pdf.cell(w, 8, txt=h, border=1, align="C")
    pdf.ln()
    total = 0
    for i, l in enumerate(lines, 1):
        pdf.cell(widths[0], 8, str(i), border=1, align="C")
        pdf.cell(widths[1], 8, str(l["품목명"])[:34], border=1)
        pdf.cell(widths[2], 8, str(l["조달상태"])[:10], border=1, align="C")
        pdf.cell(widths[3], 8, str(l["단위"]), border=1, align="C")
        pdf.cell(widths[4], 8, str(l["수량"]), border=1, align="R")
        pdf.cell(widths[5], 8, f"{l[unit_key]:,}", border=1, align="R")
        pdf.cell(widths[6], 8, f"{l[amt_key]:,}", border=1, align="R")
        pdf.ln()
        total += l[amt_key]
    pdf.set_font("ArialUnicode", size=10)
    pdf.ln(3)
    pdf.cell(0, 8, txt=f"합계 (VAT 별도): {total:,}원", ln=True, align="R")
    if note:
        pdf.set_font("ArialUnicode", size=8)
        pdf.ln(3)
        pdf.multi_cell(0, 5, txt=f"특이사항: {note}")

    os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None
    pdf.output(path)
    return path
