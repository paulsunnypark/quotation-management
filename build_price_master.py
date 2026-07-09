"""
package_price.xlsx 빌더 / 단일 가격 소스 정의

기존 CSV(기초_견적항목_테이블.csv + catalog/*.csv)를 하나의 엑셀(package_price.xlsx)로 통합한다.
이 엑셀이 프로그램의 단일 진실원본(source of truth)이 된다.
영업/관리자는 package_price.xlsx만 수정하고 앱을 새로고침하면 Track A/B 전체에 반영된다.

- `python build_price_master.py` : 현재 CSV로부터 package_price.xlsx를 (재)생성
- 시트 구성은 PRICE_SHEETS 참조
"""
import csv
import os

PRICE_XLSX = "package_price.xlsx"

# 시트명 -> 소스 CSV 경로 (엑셀 부재 시 최초 생성/복구용)
PRICE_SHEETS = {
    "base_item(A)": "기초_견적항목_테이블.csv",          # Track A
    "base_product(B)": "catalog/products.csv",            # Track B
    "registered_skus": "catalog/registered_skus.csv",
    "service_packages": "catalog/service_packages.csv",
    "package_presets": "catalog/package_presets.csv",
    "package_preset_items": "catalog/package_preset_items.csv",
    "discount_policy": "catalog/discount_policy.csv",
}

README = [
    ["package_price.xlsx — 솔루텍 견적 시스템 단일 가격 소스"],
    [""],
    ["이 파일을 수정한 뒤 앱을 새로고침(또는 재시작)하면 전체에 반영됩니다."],
    [""],
    ["시트", "용도"],
    ["base_item(A)", "Track A (레거시 단가합산) 품목·단가"],
    ["base_product(B)", "Track B 가격마스터 (계층·다단가격·식별번호·등록상태·항목코드)"],
    ["registered_skus", "조달 등록품목 (식별번호·계약종료일)"],
    ["service_packages", "공공가치 명분 패키지 10종"],
    ["package_presets", "예산대 프리셋 A~E 정의"],
    ["package_preset_items", "프리셋 구성항목"],
    ["discount_policy", "파트너 등급별 할인 한도"],
    [""],
    ["주의: 시트명과 헤더(첫 행)는 변경하지 마세요. 값/행 추가·수정만 하세요."],
]

EXTRA_PRODUCTS = [
    ["B-MOD-004", "22", "Response Module", "Response/Report Module", "수신확인·미응답 리포트 모듈", "1식", "8000000", "8000000", "4800000", "4200000", "0.60", "신규등록 후보", "Response/Report Module", "ACS/VMS 응답·리포트 add-on", "프리셋 D 구성항목에서 상품마스터로 승격", "N", "ACS/VMS"],
    ["B-IF-004", "23", "LCR Interface Bundle", "레터컬러링 PBX/I/F + User", "레터컬러링 PBX/I/F + User 조합", "식", "5000000", "5000000", "3000000", "2500000", "0.60", "등록 조합", "", "등록품목 조합 견적용", "프리셋 E 조합 항목", "N", "레터컬러링"],
    ["B-SVC-002", "24", "Service", "통합관리/연동/교육", "통합관리·연동·교육 서비스", "1식", "20000000", "20000000", "12000000", "10000000", "0.60", "커스터마이징", "", "구축 범위별 산정", "프리셋 E 커스터마이징 항목", "N", "프로젝트 범위"],
]

PRESET_ITEM_CODES = {
    "VR/STT 통합 Core (녹취내장)": "B-CORE-007",
    "IPX-VR/STT STT Channel": "B-CH-002",
    "Basic Analysis Module": "B-MOD-001",
    "설치·교육·안정화": "B-SVC-001",
    "VR/STT 분석 Core Lite (녹취제외)": "B-CORE-002",
    "VR/STT Core Standard": "B-CORE-003",
    "Evidence Management Module": "B-MOD-003",
    "IPX-SERIES IP채널라이선스": "B-CH-001",
    "Abuse Detection Module": "B-MOD-002",
    "레터컬러링 1CH": "B-CH-003",
    "설치·시나리오·연동": "B-SVC-001",
    "ACS/VMS Core Package": "B-CORE-005",
    "ACS/VMS 음성동보 1CH": "B-CH-004",
    "Response/Report Module": "B-MOD-004",
    "설치·대상자DB 연동": "B-SVC-001",
    "IPX-SERIES Core Package": "B-CORE-001",
    "레터컬러링 PBX/I/F + User": "B-IF-004",
    "전자팩스 User": "B-USER-003",
    "통합관리/연동/교육": "B-SVC-002",
}


def _read_csv(path):
    with open(path, encoding="utf-8") as f:
        return list(csv.reader(f))


def _tier_prefix(tier):
    if "Core" in tier or "Package" in tier:
        return "CORE"
    if "Channel" in tier:
        return "CH"
    if "User" in tier:
        return "USER"
    if "Integration" in tier or "Interface" in tier:
        return "IF"
    if "Service" in tier:
        return "SVC"
    return "MOD"


def _with_product_codes(rows):
    headers = rows[0]
    tier_idx = headers.index("상품계층")
    counters = {}
    out = [["항목코드"] + headers]
    for row in rows[1:]:
        prefix = _tier_prefix(row[tier_idx])
        counters[prefix] = counters.get(prefix, 0) + 1
        code = f"B-{prefix}-{counters[prefix]:03d}"
        out.append([code] + row)
    return out + EXTRA_PRODUCTS


def _with_preset_item_codes(rows):
    headers = rows[0]
    name_idx = headers.index("구성항목")
    out = [["항목코드"] + headers]
    for row in rows[1:]:
        name = row[name_idx]
        code = PRESET_ITEM_CODES.get(name)
        if not code:
            raise ValueError(f"프리셋 구성항목 코드 매핑 누락: {name}")
        out.append([code] + row)
    return out


def build(xlsx_path=PRICE_XLSX):
    """CSV들로부터 package_price.xlsx를 생성한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)

    # _README
    ws = wb.create_sheet("_README")
    for row in README:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)
    for sheet, csv_path in PRICE_SHEETS.items():
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"소스 CSV 누락: {csv_path}")
        rows = _read_csv(csv_path)
        if sheet == "base_product(B)" and rows[0][0] != "항목코드":
            rows = _with_product_codes(rows)
        if sheet == "package_preset_items" and rows[0][0] != "항목코드":
            rows = _with_preset_item_codes(rows)
        ws = wb.create_sheet(sheet)
        for r in rows:
            ws.append(r)
        # 헤더 강조
        for c in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = bold
        # 열 폭 적당히
        for c in range(1, len(rows[0]) + 1):
            ws.column_dimensions[chr(64 + c) if c <= 26 else "AA"].width = 18

    wb.save(xlsx_path)
    return xlsx_path


def read_sheet_rows(sheet_name, xlsx_path=PRICE_XLSX):
    """package_price.xlsx의 한 시트를 dict 리스트로 읽는다. 헤더=첫 행."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"시트 없음: {sheet_name} (xlsx={xlsx_path})")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [h for h in rows[0] if h is not None]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(len(headers))})
    return out


if __name__ == "__main__":
    path = build()
    print(f"[OK] {path} 생성 완료 (시트 {len(PRICE_SHEETS) + 1}개)")
